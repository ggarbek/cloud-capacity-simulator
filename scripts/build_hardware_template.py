#!/usr/bin/env python3
"""
Build public/hardware-template.xlsx — the user-facing Excel template that
mirrors the Server Builder + CPU Builder UIs 1:1 (v2.17.33).

    python3 scripts/build_hardware_template.py

Three sheets:
  1. "Servers"      — one row per server (Identity + Financial)
  2. "Server Nodes" — one row per node TYPE inside a server. Joined to
                      Servers via Group ID. Each row carries the same
                      fields the Node Type block exposes in the UI.
  3. "CPU Library"  — one row per processor.

This file is checked into the repo so Vite serves it at
/hardware-template.xlsx for the Hardware tab's "Download template" button.
The template ships with ONE sample server fully filled out so the user
has a worked example to extend.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_SHIPPED = ROOT / "public" / "hardware-template.xlsx"
OUT_TEST = ROOT / "scripts" / "test-data" / "hardware-test-data.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# Schemas — keep in sync with src/utils/hardwareTemplate.ts (parser source of
# truth) and the Server / CPU Builder UIs.
# ─────────────────────────────────────────────────────────────────────────────

SERVER_COLUMNS: list[tuple[str, str]] = [
    ("Group ID",            "Unique slug (no spaces). Becomes the persistent id Fleet Builder references. Auto-slugged from Name if you leave it blank."),
    ("Name",                "Display name shown in the Server Library (e.g. 'Gen10 R740')."),
    ("Provider",            "Free-text. Standard values used by the UI dropdown: Custom / Azure / AWS / GCP / Oracle. Anything else (e.g. 'On-prem', 'Lab') is preserved as-is and renders under the 'Custom' filter."),
    ("Cost per Rack (USD)", "Optional. Rack-level capex. The Finance tab divides this by total node count if it needs a per-node figure."),
    ("Usable Life (months)","Optional. Straight-line depreciation horizon (e.g. 60 = 5 years). Applied uniformly to every node in the rack."),
    ("Notes",               "Free-form. Shown in the Server Library card."),
]

NODE_COLUMNS: list[tuple[str, str]] = [
    ("Group ID",             "Foreign key — must match a Group ID in the 'Servers' sheet."),
    ("Node Type #",          "1, 2, 3, or 4. Up to 4 distinct node types per server. Determines visual order in the UI."),
    ("Role",                 "Compute / Utility / Other. Utility nodes are reserved for control-plane / management — the simulator never deploys VMs to them. Other behaves like Compute but carries a custom label."),
    ("Role Label",           "Only used when Role = Other. Free text (e.g. 'GPU', 'Storage', 'Inference'). Displayed as the node's tag in the UI."),
    ("Count",                "How many nodes of this type live in the rack. Required. Total nodes/rack = sum of every Count for this Group ID."),
    ("Memory / Node (GiB)",  "Required. RAM per node in this slot."),
    ("Processor",            "Optional. Should match a CPU Name from the 'CPU Library' tab (case-sensitive). Library picks auto-fill Cores/Socket + vCPUs in the UI."),
    ("Sockets / Node",       "Number of CPU sockets per node. Defaults to 2 in the UI."),
    ("Cores / Socket",       "Physical cores per CPU socket."),
    ("vCPUs / Node",         "Logical cores exposed to the OS. Usually Sockets × Cores/Socket × 2 (HT on) or × 1 (HT off)."),
    ("Network Mbps / Node",  "Per-node network bandwidth cap in Mbps. The engine treats this as a hard packing constraint."),
    ("Storage MB/s / Node",  "Per-node Premium SSD throughput cap in MB/s (bytes, not bits). 4th packing constraint."),
]

CPU_COLUMNS: list[tuple[str, str]] = [
    ("CPU Name",         "Display name. Must be unique. Example: Intel Xeon Platinum 8488C (Sapphire Rapids)."),
    ("Vendor",           "Intel / AMD / Ampere / Other. Drives default Hyperthreading (Intel = Yes, AMD/Ampere = No)."),
    ("Family",           "Architecture family. Example: Cascade Lake, Sapphire Rapids, Genoa."),
    ("Cores per Socket", "Physical cores per socket."),
    ("Hyperthreading",   "Yes or No. Yes = 2 vCPUs per physical core."),
]

# ─────────────────────────────────────────────────────────────────────────────
# Sample data — ONE worked example per the user spec (v2.17.33). The shipped
# template includes a single sample server + matching node row so users have
# a starting point. They can delete the row before adding their own.
# ─────────────────────────────────────────────────────────────────────────────

SAMPLE_SERVER = {
    "id": "sample-reference-server",
    "name": "Sample Reference Server",
    "provider": "Custom",
    "costPerRackUsd": 120000,
    "usableLifeMonths": 60,
    "notes": "Sample row — delete or edit. Demonstrates the one-server-one-node-type pattern.",
}

SAMPLE_NODES: list[dict] = [
    {
        "groupId": "sample-reference-server",
        "nodeType": 1,
        "role": "Compute",
        "roleLabel": "",
        "count": 12,
        "memoryGib": 512,
        "processor": "Intel Xeon E7-8890 v3 (Haswell)",
        "socketsPerNode": 2,
        "coresPerSocket": 18,
        "vcpusPerNode": 72,
        "networkMbpsPerNode": 50000,
        "storageThroughputMbpsPerNode": 16000,
    },
]

SAMPLE_CPUS: list[dict] = [
    # Empty by default in the shipped template — the in-app CPU Library is
    # already seeded with ~15 public-vendor CPUs at boot via publicCpuSeed.ts.
    # Users who want to author their own can fill rows in here.
]

# Test-data seed — richer payload with multiple servers + a hetero rack so
# QA / development has realistic content. Never ships to users.
TEST_SERVERS: list[dict] = [
    SAMPLE_SERVER,
    {
        "id": "compute-utility-mix",
        "name": "Compute + Utility Mix",
        "provider": "Custom",
        "costPerRackUsd": 95000,
        "usableLifeMonths": 60,
        "notes": "Worked example: a compute fleet with one utility node for control plane.",
    },
    {
        "id": "tiered-memory-rack",
        "name": "Tiered Memory Rack",
        "provider": "Custom",
        "costPerRackUsd": 180000,
        "usableLifeMonths": 72,
        "notes": "Mixed-memory rack — small + large nodes in the same chassis.",
    },
]

TEST_NODES: list[dict] = [
    *SAMPLE_NODES,
    # compute-utility-mix: 11 compute + 1 utility
    {
        "groupId": "compute-utility-mix",
        "nodeType": 1,
        "role": "Compute",
        "roleLabel": "",
        "count": 11,
        "memoryGib": 512,
        "processor": "Intel Xeon Platinum 8488C (Sapphire Rapids)",
        "socketsPerNode": 2,
        "coresPerSocket": 48,
        "vcpusPerNode": 192,
        "networkMbpsPerNode": 100000,
        "storageThroughputMbpsPerNode": 16000,
    },
    {
        "groupId": "compute-utility-mix",
        "nodeType": 2,
        "role": "Utility",
        "roleLabel": "",
        "count": 1,
        "memoryGib": 128,
        "processor": "Intel Xeon Platinum 8488C (Sapphire Rapids)",
        "socketsPerNode": 2,
        "coresPerSocket": 16,
        "vcpusPerNode": 32,
        "networkMbpsPerNode": 25000,
        "storageThroughputMbpsPerNode": 8000,
    },
    # tiered-memory-rack: 6 × 1024 GiB + 6 × 4096 GiB
    {
        "groupId": "tiered-memory-rack",
        "nodeType": 1,
        "role": "Compute",
        "roleLabel": "",
        "count": 6,
        "memoryGib": 1024,
        "processor": "AMD EPYC 9654 (Genoa)",
        "socketsPerNode": 2,
        "coresPerSocket": 96,
        "vcpusPerNode": 192,
        "networkMbpsPerNode": 100000,
        "storageThroughputMbpsPerNode": 16000,
    },
    {
        "groupId": "tiered-memory-rack",
        "nodeType": 2,
        "role": "Compute",
        "roleLabel": "",
        "count": 6,
        "memoryGib": 4096,
        "processor": "AMD EPYC 9654 (Genoa)",
        "socketsPerNode": 2,
        "coresPerSocket": 96,
        "vcpusPerNode": 192,
        "networkMbpsPerNode": 100000,
        "storageThroughputMbpsPerNode": 16000,
    },
]

TEST_CPUS: list[dict] = [
    {"name": "Intel Xeon E7-8890 v3 (Haswell)",     "vendor": "Intel", "family": "Haswell-EX",      "coresPerSocket": 18,  "hyperthreading": True},
    {"name": "Intel Xeon Platinum 8488C (Sapphire Rapids)", "vendor": "Intel", "family": "Sapphire Rapids", "coresPerSocket": 48, "hyperthreading": True},
    {"name": "AMD EPYC 9654 (Genoa)",               "vendor": "AMD",   "family": "Genoa",            "coresPerSocket": 96,  "hyperthreading": False},
]

# ─────────────────────────────────────────────────────────────────────────────
# Styling — restrained; this is a data file, not a presentation.
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(bold=True, color="A7F3D0", size=11, name="SF Pro Text")
TITLE_FONT = Font(bold=True, color="22C55E", size=14, name="SF Pro Display")
HINT_FONT = Font(italic=True, color="64748B", size=10, name="SF Pro Text")
DATA_FONT = Font(color="0F172A", size=11, name="SF Mono")
EDGE = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(top=EDGE, bottom=EDGE, left=EDGE, right=EDGE)


def style_header_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER


def write_tab(ws, title: str, columns: list[tuple[str, str]], rows: list[list]) -> None:
    """Layout: row 1 title, row 2 hint, row 4 column headers, row 5+ data."""
    n_cols = len(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    hint = (
        "Each row = one entry. Hover any column header for what to put. "
        "Blank rows and rows missing required fields are skipped on import."
    )
    ws.cell(row=2, column=1, value=hint).font = HINT_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 4 — column headers with cell comments for guidance
    for c, (header, hint_text) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=c, value=header)
        cell.comment = Comment(hint_text, "Capacity Simulator")
        cell.comment.width = 320
        cell.comment.height = 110
    style_header_row(ws, 4, n_cols)
    ws.row_dimensions[4].height = 32

    # Data
    for r_offset, row in enumerate(rows, start=5):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r_offset, column=c, value=value)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = CELL_BORDER

    # Column widths — heuristic per column type
    for c, (header, _) in enumerate(columns, start=1):
        width = max(len(header) + 2, 14)
        if "Notes" in header or "Processor" in header or "Family" in header:
            width = 36
        elif "Name" in header:
            width = 32
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = "A5"


def yes_no(v) -> str:
    return "Yes" if v else "No"


def _server_rows(seed: list[dict]) -> list[list]:
    return [
        [
            s["id"],
            s["name"],
            s.get("provider", ""),
            s.get("costPerRackUsd", ""),
            s.get("usableLifeMonths", ""),
            s.get("notes", ""),
        ]
        for s in seed
    ]


def _node_rows(seed: list[dict]) -> list[list]:
    return [
        [
            n["groupId"],
            n["nodeType"],
            n["role"],
            n.get("roleLabel", ""),
            n["count"],
            n["memoryGib"],
            n.get("processor", ""),
            n.get("socketsPerNode", ""),
            n.get("coresPerSocket", ""),
            n.get("vcpusPerNode", ""),
            n.get("networkMbpsPerNode", ""),
            n.get("storageThroughputMbpsPerNode", ""),
        ]
        for n in seed
    ]


def _cpu_rows(seed: list[dict]) -> list[list]:
    return [
        [
            c["name"],
            c.get("vendor", ""),
            c.get("family", ""),
            c.get("coresPerSocket", ""),
            yes_no(c.get("hyperthreading", False)),
        ]
        for c in seed
    ]


def _write_workbook(
    path: Path,
    servers: list[list],
    nodes: list[list],
    cpus: list[list],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_servers = wb.active
    ws_servers.title = "Servers"
    write_tab(ws_servers, "Servers", SERVER_COLUMNS, servers)

    ws_nodes = wb.create_sheet("Server Nodes")
    write_tab(ws_nodes, "Server Nodes", NODE_COLUMNS, nodes)

    ws_cpu = wb.create_sheet("CPU Library")
    write_tab(ws_cpu, "CPU Library", CPU_COLUMNS, cpus)

    wb.save(path)


def main() -> None:
    # ─── Shipped template (public/) — one sample server + matching node row.
    _write_workbook(
        OUT_SHIPPED,
        servers=_server_rows([SAMPLE_SERVER]),
        nodes=_node_rows(SAMPLE_NODES),
        cpus=_cpu_rows(SAMPLE_CPUS),
    )
    print(f"Wrote {OUT_SHIPPED.relative_to(ROOT)} — shipped template (1 sample server).")

    # ─── Test-data file (scripts/test-data/) — richer seed for QA.
    _write_workbook(
        OUT_TEST,
        servers=_server_rows(TEST_SERVERS),
        nodes=_node_rows(TEST_NODES),
        cpus=_cpu_rows(TEST_CPUS),
    )
    print(
        f"Wrote {OUT_TEST.relative_to(ROOT)} — test data "
        f"({len(TEST_SERVERS)} servers, {len(TEST_NODES)} node-types, {len(TEST_CPUS)} CPUs)."
    )


if __name__ == "__main__":
    main()
