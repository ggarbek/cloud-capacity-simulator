#!/usr/bin/env python3
"""
Build public/bom-template.xlsx — the user-facing Excel template for uploading
a Bill of Materials. v2.4.3 addition; mirrors the layout idiom of
hardware-template.xlsx / vm-template.xlsx so the three feel consistent.

    python3 scripts/build_bom_template.py

Schema (single tab, source-of-truth lives in lockstep with
src/utils/bomTemplate.ts — if you change one, change the other):

    | VM Size Name | Quantity | Notes |

The "VM Size Name" cell must match an entry in the user's uploaded VM catalog
(state.userVms). Rows whose VM Size Name doesn't resolve are surfaced as
unknowns in the BOM (today's existing "not in catalog" banner).

The template ships **with a few illustrative seed rows** so the user has a
concrete worked example — but those rows reference vendor-public SKU names
only (no proprietary data per the Decoupling Doctrine).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "public" / "bom-template.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# Schema
# ─────────────────────────────────────────────────────────────────────────────

BOM_COLUMNS = [
    ("VM Size Name", "Vendor SKU name. Must match a row uploaded in the VMs tab. Example: Standard_M64s_v2, m7i.4xlarge, n2-highmem-32."),
    ("Quantity",     "How many of this VM to include in the BOM. Integer ≥ 1."),
    ("Notes",        "Optional free-text — surfaced in the BOM row hover. Ignored by the engine."),
]

# Seed rows — a tiny but representative multi-cloud mix using only
# vendor-published SKU names. Decoupling doctrine: never include real
# customer BoMs here; rows must be derivable from the public catalog.
SEED_ROWS = [
    ("Standard_M64s_v2",      4, "Example: Azure Mv2 baseline"),
    ("Standard_M128ms_v2",    2, "Example: Azure Mv2 high-mem"),
    ("Standard_M832is_16_v3", 1, "Example: Azure Mv3 HM"),
    ("m7i.4xlarge",           4, "Example: AWS m7i general"),
    ("n2-highmem-32",         2, "Example: GCP n2 highmem"),
]

# Visual styling — restrained, matches the other templates.
HEADER_FILL  = PatternFill("solid", fgColor="0F172A")  # slate-900
HEADER_FONT  = Font(bold=True, color="A7F3D0", size=11, name="SF Pro Text")
TITLE_FONT   = Font(bold=True, color="22C55E", size=14, name="SF Pro Display")
HINT_FONT    = Font(italic=True, color="64748B", size=10, name="SF Pro Text")
DATA_FONT    = Font(color="0F172A", size=11, name="SF Mono")
EDGE         = Side(style="thin", color="CBD5E1")
CELL_BORDER  = Border(top=EDGE, bottom=EDGE, left=EDGE, right=EDGE)


def style_header_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER


def write_tab(ws, title: str, columns, rows) -> None:
    """Layout: row 1 title, row 2 hint, row 4 column headers, row 5+ data."""
    n_cols = len(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    hint = (
        "Each row = one BOM line. VM Size Name must match a row uploaded in the VMs tab. "
        "Blank rows and rows missing required fields are skipped on import."
    )
    ws.cell(row=2, column=1, value=hint).font = HINT_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Column headers + hover-comment hints.
    for c, (header, hint_text) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=c, value=header)
        cell.comment = Comment(hint_text, "Capacity Simulator")
        cell.comment.width = 360
        cell.comment.height = 110
    style_header_row(ws, 4, n_cols)
    ws.row_dimensions[4].height = 32

    for r_offset, row in enumerate(rows, start=5):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r_offset, column=c, value=value)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = CELL_BORDER

    # Column widths — heuristic per column type.
    for c, (header, _) in enumerate(columns, start=1):
        width = max(len(header) + 4, 18)
        if "Name" in header:
            width = 32
        elif "Notes" in header:
            width = 40
        elif "Quantity" in header:
            width = 14
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = "A5"


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    write_tab(ws, "Bill of Materials", BOM_COLUMNS, SEED_ROWS)

    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(ROOT)} — {len(SEED_ROWS)} seed row(s).")


if __name__ == "__main__":
    main()
